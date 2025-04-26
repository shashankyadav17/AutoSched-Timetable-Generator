import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Load generated timetable JSON
with open("generated_timetable_all.json") as f:
    data = json.load(f)

# Load general settings (needed for Break and timings)
with open("timetable_input_all_semesters.json") as f:
    settings_data = json.load(f)

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

def time_to_minutes(t):
    parts = t.split(':')
    return int(parts[0]) * 60 + int(parts[1])

def minutes_to_time(m):
    h = m // 60
    mm = m % 60
    suffix = "AM"
    if h >= 12:
        suffix = "PM"
        if h > 12:
            h -= 12
    return f"{h}:{mm:02d} {suffix}"

output_dir = "output"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

output_file = os.path.join(output_dir, "generated_timetable.xlsx")

for sem_key, sections in data.items():
    sem_id = int(sem_key.replace('sem_', ''))
    sem_settings = settings_data['semesters'][sem_key]['general_settings']

    working_days = sem_settings['working_days']
    periods_per_day = sem_settings['periods_per_day']
    break_after = sem_settings['break_after_period']
    break_duration = sem_settings['break_duration']
    start_time_str = sem_settings['start_time']
    period_duration = sem_settings['period_duration']

    start_minutes = time_to_minutes(start_time_str)

    for section, timetable in sections.items():
        sheet_name = f"{sem_id}-{section}"
        ws = wb.create_sheet(title=sheet_name)

        # Create Header Row with period timings and break
        header = ['Day / Hour']
        for i in range(1, periods_per_day + 1):
            if i == break_after + 1:
                # Break timing
                break_start = start_minutes + period_duration * break_after
                break_end = break_start + break_duration
                header.append(f"Break\n{minutes_to_time(break_start)} - {minutes_to_time(break_end)}")
            # Period timing
            period_start = start_minutes + period_duration * (i - 1)
            # Add break duration after break period
            if i > break_after:
                period_start += break_duration
            period_end = period_start + period_duration
            header.append(f"Period {i}\n{minutes_to_time(period_start)} - {minutes_to_time(period_end)}")
        ws.append(header)

        # Style header
        for col in ws[1]:
            col.font = Font(bold=True)
            col.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Fill timetable rows for each day
        for day in working_days:
            row = [day]
            slots = timetable.get(day, [""] * periods_per_day)
            break_inserted = False
            for idx, slot in enumerate(slots):
                period_num = idx + 1
                if period_num == break_after + 1 and not break_inserted:
                    row.append("Break")
                    break_inserted = True
                row.append(slot)
            ws.append(row)

# Save
wb.save(output_file)
print(f"Excel timetable saved as {output_file}")
